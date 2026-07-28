# -*- coding: utf-8 -*-
"""
Экспорт данных бизнеса в CSV, Excel и PDF.

Логика разделена на два слоя:
  1. dataset(bid, key) -> {"key","title","columns","rows"} — что выгружаем (из базы);
  2. to_csv / to_xlsx / to_pdf — как это упаковать в файл.
Чтобы добавить новый набор данных, достаточно вписать сборщик в DATASETS —
форматы подхватят его сами, без правок writer'ов.
"""
import csv
import io
import json
import datetime

import database

# наборы данных, которые может выгрузить владелец
DATASET_KEYS = ["clients", "finance", "documents", "reports", "analytics", "goals"]
DATASET_TITLES = {
    "clients":   "Клиенты",
    "finance":   "Финансы",
    "documents": "Документы",
    "reports":   "Отчёты",
    "analytics": "Аналитика",
    "goals":     "Цели",
}
FORMATS = ["csv", "xlsx", "pdf"]


def _day(s):
    return (str(s or "")[:10]).strip()


# ---------- СБОРЩИКИ НАБОРОВ ----------

def _clients(bid):
    rows = []
    for c in database.list_clients(bid, limit=100000):
        rows.append([
            c.get("name") or "", c.get("phone") or "", c.get("birthday") or "",
            c.get("favorite") or "", c.get("orders_count") or 0,
            int(c.get("total_spent") or 0), (c.get("notes") or "").strip(),
        ])
    return ["Имя", "Телефон", "День рождения", "Предпочтения",
            "Заказов", "Сумма покупок, ₽", "Заметки"], rows


def _finance(bid):
    rows = []
    for e in database.all_finance_entries(bid):
        rows.append([
            _day(e.get("op_date") or e.get("created_at")),
            "Доход" if e.get("kind") == "income" else "Расход",
            e.get("category") or "", int(e.get("amount") or 0),
            e.get("counterparty") or "", (e.get("note") or "").strip(),
        ])
    return ["Дата", "Тип", "Категория", "Сумма, ₽", "Контрагент", "Назначение"], rows


def _documents(bid):
    rows = []
    for d in database.list_documents(bid):
        rows.append([d.get("filename") or "", d.get("chunks") or 0,
                     _day(d.get("created_at"))])
    return ["Название", "Фрагментов", "Загружен"], rows


def _reports(bid):
    """Отчёты — еженедельные обзоры бизнеса, разложенные по колонкам."""
    rows = []
    for r in database.list_weekly_reviews(bid, limit=200):
        try:
            p = json.loads(r.get("payload") or "{}")
        except json.JSONDecodeError:
            p = {}
        fin = p.get("finance") or {}
        rows.append([
            r.get("week_start") or "", int(fin.get("income") or 0),
            int(fin.get("expense") or 0), int(fin.get("profit") or 0),
            (p.get("achievements") or "").strip(), (p.get("mistakes") or "").strip(),
            (p.get("next_week") or "").strip(),
        ])
    return ["Неделя (с)", "Доход, ₽", "Расход, ₽", "Прибыль, ₽",
            "Достижения", "Ошибки", "План на следующую неделю"], rows


def _analytics(bid):
    """Аналитика — сводные показатели «показатель → значение»."""
    stats = database.business_stats(bid)
    fin = database.finance_summary(bid)
    sig = database.growth_signals(bid)
    pairs = [
        ("Клиентов в базе", stats.get("clients", 0)),
        ("Обработано сообщений", stats.get("messages", 0)),
        ("Заказов всего", stats.get("orders_total", 0)),
        ("Заказов выполнено", stats.get("orders_done", 0)),
        ("Спящих клиентов (30+ дней)", sig.get("sleeping", 0)),
        ("Клиентов с повторными заказами", sig.get("repeat_clients", 0)),
        ("Доход всего, ₽", int(fin.get("income") or 0)),
        ("Расход всего, ₽", int(fin.get("expense") or 0)),
        ("Прибыль, ₽", int(fin.get("profit") or 0)),
        ("Маржа, %", sig.get("margin") if sig.get("margin") is not None else "—"),
    ]
    for c in (fin.get("by_category") or [])[:10]:
        kind = "доход" if c.get("kind") == "income" else "расход"
        pairs.append((f"По категории «{c.get('category')}» ({kind}), ₽", int(c.get("total") or 0)))
    return ["Показатель", "Значение"], [[k, v] for k, v in pairs]


def _goals(bid):
    rows = []
    status_ru = {"active": "в работе", "done": "достигнута", "dropped": "снята"}
    for g in database.list_goals(bid):
        rows.append([
            g.get("title") or "", g.get("metric_name") or g.get("metric") or "",
            g.get("target") or 0, g.get("current") or 0,
            f"{g.get('percent') or 0}%", g.get("deadline") or "—",
            status_ru.get(g.get("status"), g.get("status") or ""),
        ])
    return ["Цель", "Метрика", "Цель (число)", "Сейчас", "Прогресс", "Срок", "Статус"], rows


_BUILDERS = {
    "clients": _clients, "finance": _finance, "documents": _documents,
    "reports": _reports, "analytics": _analytics, "goals": _goals,
}


def dataset(bid, key):
    """Собрать набор данных: {key, title, columns, rows}."""
    if key not in _BUILDERS:
        return None
    columns, rows = _BUILDERS[key](bid)
    return {"key": key, "title": DATASET_TITLES[key], "columns": columns, "rows": rows}


# ---------- ФОРМАТЫ ----------

def to_csv(ds) -> bytes:
    buf = io.StringIO()
    # ';' и BOM — чтобы русский Excel открыл кириллицу и колонки без плясок с бубном
    w = csv.writer(buf, delimiter=";", quoting=csv.QUOTE_MINIMAL)
    w.writerow(ds["columns"])
    for row in ds["rows"]:
        w.writerow(["" if v is None else v for v in row])
    return b"\xef\xbb\xbf" + buf.getvalue().encode("utf-8")


def to_xlsx(ds) -> bytes:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill

    wb = openpyxl.Workbook()
    sh = wb.active
    sh.title = ds["title"][:31]

    head_font = Font(bold=True, color="FFFFFF")
    head_fill = PatternFill("solid", fgColor="8052FF")
    sh.append(ds["columns"])
    for cell in sh[1]:
        cell.font = head_font
        cell.fill = head_fill
        cell.alignment = Alignment(vertical="center", wrap_text=True)

    for row in ds["rows"]:
        sh.append(["" if v is None else v for v in row])

    # ширина колонок по содержимому (с потолком, чтобы заметки не растягивали лист)
    for i, col in enumerate(ds["columns"], start=1):
        longest = max([len(str(col))] + [len(str(r[i - 1])) for r in ds["rows"]] or [0])
        sh.column_dimensions[openpyxl.utils.get_column_letter(i)].width = min(60, max(12, longest + 2))
    sh.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()


_PDF_FONT = None


def _pdf_font():
    """Зарегистрировать шрифт с кириллицей (иначе reportlab рисует пустые квадраты)."""
    global _PDF_FONT
    if _PDF_FONT:
        return _PDF_FONT
    from reportlab.pdfbase import pdfmetrics
    from reportlab.pdfbase.ttfonts import TTFont
    for name, path in [("Arial", r"C:\Windows\Fonts\arial.ttf"),
                       ("ArialB", r"C:\Windows\Fonts\arialbd.ttf"),
                       ("DejaVu", "/usr/share/fonts/truetype/dejavu/DejaVuSans.ttf")]:
        try:
            pdfmetrics.registerFont(TTFont(name, path))
            if _PDF_FONT is None:
                _PDF_FONT = name
        except Exception:
            pass
    return _PDF_FONT or "Helvetica"


def to_pdf(ds) -> bytes:
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.lib import colors
    from reportlab.lib.units import mm
    from reportlab.lib.styles import ParagraphStyle
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer

    base = _pdf_font()
    bold = "ArialB" if base == "Arial" else base
    cell_style = ParagraphStyle("cell", fontName=base, fontSize=8, leading=10)
    head_style = ParagraphStyle("head", fontName=bold, fontSize=8, leading=10, textColor=colors.white)
    title_style = ParagraphStyle("title", fontName=bold, fontSize=15, leading=18)
    meta_style = ParagraphStyle("meta", fontName=base, fontSize=9, textColor=colors.grey)

    def cell(v, style):
        return Paragraph(str("" if v is None else v).replace("\n", "<br/>")
                         .replace("&", "&amp;").replace("<", "&lt;").replace(">", "&gt;"), style)

    data = [[cell(c, head_style) for c in ds["columns"]]]
    for row in ds["rows"]:
        data.append([cell(v, cell_style) for v in row])

    out = io.BytesIO()
    doc = SimpleDocTemplate(out, pagesize=landscape(A4),
                            leftMargin=14 * mm, rightMargin=14 * mm,
                            topMargin=14 * mm, bottomMargin=14 * mm)
    today = datetime.date.today().strftime("%d.%m.%Y")
    elems = [
        Paragraph(f"VELOR AI · {ds['title']}", title_style),
        Paragraph(f"Выгружено {today} · строк: {len(ds['rows'])}", meta_style),
        Spacer(1, 8),
    ]
    if ds["rows"]:
        tbl = Table(data, repeatRows=1)
        tbl.setStyle(TableStyle([
            ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#8052FF")),
            ("GRID", (0, 0), (-1, -1), 0.4, colors.HexColor("#CCCCCC")),
            ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F4F1FF")]),
            ("VALIGN", (0, 0), (-1, -1), "TOP"),
            ("LEFTPADDING", (0, 0), (-1, -1), 5), ("RIGHTPADDING", (0, 0), (-1, -1), 5),
            ("TOPPADDING", (0, 0), (-1, -1), 4), ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
        ]))
        elems.append(tbl)
    else:
        elems.append(Paragraph("Данных для выгрузки пока нет.", meta_style))
    doc.build(elems)
    return out.getvalue()


_WRITERS = {"csv": to_csv, "xlsx": to_xlsx, "pdf": to_pdf}
_MIME = {
    "csv": "text/csv; charset=utf-8",
    "xlsx": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    "pdf": "application/pdf",
}


def export(bid, key, fmt):
    """Собрать набор и упаковать в формат. Возвращает (bytes, mime, filename)."""
    if key not in _BUILDERS or fmt not in _WRITERS:
        return None
    ds = dataset(bid, key)
    blob = _WRITERS[fmt](ds)
    today = datetime.date.today().isoformat()
    filename = f"velor-{key}-{today}.{fmt}"
    return blob, _MIME[fmt], filename
