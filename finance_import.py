"""
Импорт финансов: банковская выписка → операции в базе.

Конвейер намеренно разбит на три независимых шага:

    ИСТОЧНИК  →  нормализованные операции  →  КАТЕГОРИЗАТОР  →  сохранение

Источник — это функция, которая умеет достать операции откуда угодно (файл CSV,
XLSX, PDF, а завтра — API банка, 1С или CRM) и вернуть их в одном общем виде:

    {"date": "2026-07-20", "amount": 24000, "direction": "income"|"expense",
     "description": "Оплата по счёту 41", "counterparty": "ООО Ромашка",
     "external_id": "строка, по которой узнаём дубль"}

Всё, что дальше — распознавание доходов/расходов, категории, дедупликация,
запись в БД — уже не знает, откуда пришли данные. Поэтому новая интеграция
добавляется одной функцией и одной строкой в SOURCES, без правок остального.
"""
import csv
import datetime
import io
import re

# ============================================================
#  Категории расходов
# ============================================================

CATEGORIES = ["аренда", "реклама", "доставка", "закупки", "зарплата", "налоги", "прочее"]

# Ключевые слова → категория. Обычные правила покрывают большинство операций,
# ИИ зовём только на то, что здесь не нашлось.
RULES = {
    "аренда":   ["аренд", "аренда помещ", "субаренд", "коммунальн", "жкх", "электроэнерг"],
    "реклама":  ["реклам", "яндекс директ", "директ", "vk ads", "таргет", "маркетинг",
                 "продвижен", "avito", "авито", "google ads", "смм"],
    "доставка": ["доставк", "логистик", "курьер", "сдэк", "cdek", "boxberry", "почта россии",
                 "деловые линии", "транспортн", "перевозк", "такси"],
    "закупки":  ["закуп", "поставщик", "товар", "материал", "сырь", "опт", "склад",
                 "оптов", "комплектующ"],
    "зарплата": ["зарплат", "заработн", "аванс сотруд", "оплата труда", "премия",
                 "отпускн", "больничн", "фот"],
    "налоги":   ["налог", "ндс", "усн", "ндфл", "страхов взнос", "пенсионн", "фнс",
                 "госпошлин", "взнос в фонд", "торговый сбор"],
}

# Признаки того, что операция — доход, даже если знак суммы потерялся
INCOME_HINTS = ["оплата от", "поступлен", "выручка", "возврат от", "зачислен",
                "перевод от", "продажа", "эквайринг", "пополнен"]


def _norm(text):
    return re.sub(r"\s+", " ", (text or "")).strip().lower()


def guess_category(description, counterparty="", learned=None):
    """
    Категория расхода по тексту операции.

    Возвращает (категория, уверенность, откуда): откуда — 'learned' | 'rules' | None.
    None в категории означает «правилами не понял, нужен ИИ».
    Выученные правила проверяем первыми: выбор владельца важнее наших умолчаний.
    """
    text = _norm(description + " " + counterparty)
    if not text:
        return None, 0.0, None

    for pattern, category in (learned or {}).items():
        if pattern and pattern in text:
            return category, 1.0, "learned"

    for category, words in RULES.items():
        for w in words:
            if w in text:
                return category, 0.9, "rules"

    return None, 0.0, None


# слова, по которым нельзя узнать операцию — в правило их не берём
_STOP = {"оплата", "платеж", "платёж", "по", "счет", "счёт", "за", "от", "для", "договор",
         "договору", "ндс", "без", "руб", "рублей", "сумма", "перевод", "номер", "на"}


def learn_pattern(description, counterparty=""):
    """
    По какому куску текста узнавать похожие операции в будущем.

    Контрагент надёжнее всего: «ООО Ромашка» будет в каждой такой строке.
    Если его нет — берём из назначения два самых значимых слова, выбрасывая
    номера счетов, даты и канцелярит, иначе правило поймает пол-выписки.
    """
    party = _norm(counterparty)
    if len(party) >= 3:
        return party[:60]

    words = [w for w in re.findall(r"[а-яёa-z]{3,}", _norm(description)) if w not in _STOP]
    if not words:
        return None
    return " ".join(words[:2])[:60]


def looks_like_income(description, counterparty=""):
    text = _norm(description + " " + counterparty)
    return any(h in text for h in INCOME_HINTS)


# ============================================================
#  Источники данных
# ============================================================

def _to_amount(value):
    """Сумма из любой каши: '−1 234,56 ₽', '(1234.56)', '1 234'."""
    if value is None:
        return None
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s:
        return None
    negative = s.startswith("(") and s.endswith(")")
    s = s.replace(" ", " ").replace("−", "-").replace("–", "-")
    s = re.sub(r"[^\d,.\-]", "", s)
    if not s or s in "-.,":
        return None
    # 1 234,56 → 1234.56;  1,234.56 → 1234.56
    if "," in s and "." in s:
        s = s.replace(",", "") if s.rfind(".") > s.rfind(",") else s.replace(".", "").replace(",", ".")
    else:
        s = s.replace(",", ".")
    try:
        amount = float(s)
    except ValueError:
        return None
    return -amount if negative else amount


def _to_date(value):
    """Дата в ISO из типичных форматов выписок."""
    if isinstance(value, datetime.datetime):
        return value.date().isoformat()
    if isinstance(value, datetime.date):
        return value.isoformat()
    s = str(value or "").strip()[:19]
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%Y/%m/%d", "%d-%m-%Y",
                "%Y-%m-%d %H:%M:%S", "%d.%m.%Y %H:%M:%S", "%d.%m.%y"):
        try:
            return datetime.datetime.strptime(s, fmt).date().isoformat()
        except ValueError:
            continue
    m = re.search(r"(\d{2})[.\-/](\d{2})[.\-/](\d{4})", s)
    if m:
        return f"{m.group(3)}-{m.group(2)}-{m.group(1)}"
    return None


# как называются нужные колонки в выписках разных банков
COLUMNS = {
    "date":         ["дата", "дата операции", "дата платежа", "дата проводки", "date",
                     "дата и время", "операция дата"],
    "description":  ["назначение платежа", "назначение", "описание", "комментарий",
                     "детали", "description", "примечание", "содержание операции"],
    "counterparty": ["контрагент", "получатель", "плательщик", "наименование контрагента",
                     "payee", "корреспондент", "отправитель"],
    "amount":       ["сумма", "сумма операции", "сумма в валюте счёта", "amount",
                     "сумма платежа", "сумма руб"],
    "income":       ["приход", "поступление", "кредит", "зачисление", "доход", "credit"],
    "expense":      ["расход", "списание", "дебет", "списано", "debit"],
}


def _match_column(header):
    """К какому смыслу относится колонка выписки."""
    h = _norm(header)
    if not h:
        return None
    for key, names in COLUMNS.items():
        for n in names:
            if h == n or h.startswith(n) or n in h:
                return key
    return None


def _rows_to_operations(rows):
    """
    Строки таблицы (первая — заголовок) → нормализованные операции.
    Понимает оба формата: одна колонка «сумма» со знаком и две колонки приход/расход.
    """
    if not rows:
        return []

    # заголовок ищем в первых пяти строках: банки любят шапку с реквизитами
    header_at, mapping = None, {}
    for i, row in enumerate(rows[:5]):
        found = {}
        for j, cell in enumerate(row):
            key = _match_column(cell)
            if key and key not in found:
                found[key] = j
        if "date" in found and ("amount" in found or "income" in found or "expense" in found):
            header_at, mapping = i, found
            break
    if header_at is None:
        return []

    operations = []
    for row in rows[header_at + 1:]:
        if not any(str(c or "").strip() for c in row):
            continue
        get = lambda key: (row[mapping[key]] if key in mapping and mapping[key] < len(row) else None)

        date = _to_date(get("date"))
        if not date:
            continue

        description = str(get("description") or "").strip()
        counterparty = str(get("counterparty") or "").strip()

        income = _to_amount(get("income"))
        expense = _to_amount(get("expense"))
        if income or expense:
            if income and income > 0:
                amount, direction = income, "income"
            elif expense and expense != 0:
                amount, direction = abs(expense), "expense"
            else:
                continue
        else:
            # одна колонка суммы: знак и есть направление
            value = _to_amount(get("amount"))
            if value is None or value == 0:
                continue
            amount = abs(value)
            direction = "income" if value > 0 else "expense"

        operations.append({
            "date": date,
            "amount": int(round(amount)),
            "direction": direction,
            "description": description[:400],
            "counterparty": counterparty[:200],
            "external_id": f"{date}|{int(round(amount))}|{_norm(description)[:60]}",
        })
    return operations


def from_csv(raw: bytes):
    """CSV-выписка. Кодировку и разделитель определяем сами."""
    text = None
    for encoding in ("utf-8-sig", "cp1251", "utf-8", "latin-1"):
        try:
            text = raw.decode(encoding)
            break
        except UnicodeDecodeError:
            continue
    if text is None:
        return []
    sample = text[:4000]
    delimiter = max([",", ";", "\t", "|"], key=lambda d: sample.count(d))
    rows = list(csv.reader(io.StringIO(text), delimiter=delimiter))
    return _rows_to_operations(rows)


def from_xlsx(raw: bytes):
    """XLSX-выписка — читаем первый лист."""
    try:
        import openpyxl
    except ImportError:
        return []
    book = openpyxl.load_workbook(io.BytesIO(raw), read_only=True, data_only=True)
    rows = [list(r) for r in book[book.sheetnames[0]].iter_rows(values_only=True, max_row=3000)]
    book.close()
    return _rows_to_operations(rows)


# Строка PDF-выписки: дата, назначение, сумма в конце.
# Разряды суммы — строго по три цифры, иначе номер счёта из назначения
# («по счету 41 24 000,00») прилипает к сумме и делает из неё 4 124 000.
_PDF_LINE = re.compile(
    r"(\d{2}[.\-/]\d{2}[.\-/]\d{2,4})\s+(.+?)\s+"
    r"([-−(]?\s?\d{1,3}(?:[\s ]\d{3})*[.,]\d{2}\)?)\s*$")


def from_pdf(raw: bytes):
    """
    PDF-выписка. Таблиц в PDF нет — есть строки текста, поэтому разбираем
    построчно: дата в начале, сумма в конце, между ними назначение.
    """
    try:
        from pypdf import PdfReader
    except ImportError:
        return []
    try:
        reader = PdfReader(io.BytesIO(raw))
        text = "\n".join((page.extract_text() or "") for page in reader.pages[:40])
    except Exception:
        return []

    operations = []
    for line in text.splitlines():
        m = _PDF_LINE.match(line.strip())
        if not m:
            continue
        date = _to_date(m.group(1))
        amount = _to_amount(m.group(3))
        if not date or not amount:
            continue
        description = m.group(2).strip()
        # В PDF знак часто теряется, поэтому минус — расход наверняка,
        # а плюс считаем доходом только при словах вроде «оплата от», «поступление».
        direction = "expense" if amount < 0 else (
            "income" if looks_like_income(description) else "expense")
        operations.append({
            "date": date,
            "amount": int(round(abs(amount))),
            "direction": direction,
            "description": description[:400],
            "counterparty": "",
            "external_id": f"{date}|{int(round(abs(amount)))}|{_norm(description)[:60]}",
        })
    return operations


# Реестр источников. Новая интеграция (банк по API, 1С, CRM) добавляется сюда
# одной строкой — остальной конвейер её даже не заметит.
SOURCES = {
    "csv":  from_csv,
    "xlsx": from_xlsx,
    "xls":  from_xlsx,
    "pdf":  from_pdf,
}


def parse(filename: str, raw: bytes):
    """Разобрать файл выписки по расширению. Возвращает список операций."""
    ext = (filename or "").rsplit(".", 1)[-1].lower()
    source = SOURCES.get(ext)
    return source(raw) if source else []
